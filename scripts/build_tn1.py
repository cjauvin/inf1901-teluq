from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
import os

# Racine du dépôt (parent du dossier scripts/), pour y écrire le .xlsx
_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

FONT = "Arial"
HDR_FILL = PatternFill("solid", fgColor="1F4E78")
HDR_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name=FONT, bold=True, size=14, color="1F4E78")
BASE_FONT = Font(name=FONT, size=11)
GREEN = PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid")
GREEN_OK = PatternFill(start_color="FFA9D08E", end_color="FFA9D08E", fill_type="solid")
AMBER = PatternFill(start_color="FFFFE699", end_color="FFFFE699", fill_type="solid")
RED = PatternFill(start_color="FFF4CCCC", end_color="FFF4CCCC", fill_type="solid")
GREYFILL = PatternFill(start_color="FFF2F2F2", end_color="FFF2F2F2", fill_type="solid")
GREYFONT = Font(name=FONT, size=11, color="808080")
DERIV_FILL = PatternFill("solid", fgColor="FFF2CC")
ESP_FILL = PatternFill("solid", fgColor="DDEBF7")
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()

# ---------- Faits ----------
fa = wb.active
fa.title = "Faits"
facts = [
    ("a du poil", "observable"),
    ("allaite ses petits", "observable"),
    ("a des plumes", "observable"),
    ("pond des œufs", "observable"),
    ("mange de la viande", "observable"),
    ("a des dents pointues", "observable"),
    ("a des griffes", "observable"),
    ("a des sabots", "observable"),
    ("rumine", "observable"),
    ("robe fauve", "observable"),
    ("taches sombres", "observable"),
    ("rayures noires", "observable"),
    ("long cou", "observable"),
    ("incapable de voler", "observable"),
    ("nage", "observable"),
    ("plumage noir et blanc", "observable"),
    ("mammifère", "catégorie (déduite)"),
    ("oiseau", "catégorie (déduite)"),
    ("carnivore", "catégorie (déduite)"),
    ("guépard", "espèce (conclusion)"),
    ("tigre", "espèce (conclusion)"),
    ("girafe", "espèce (conclusion)"),
    ("zèbre", "espèce (conclusion)"),
    ("autruche", "espèce (conclusion)"),
    ("manchot", "espèce (conclusion)"),
]
fa["A1"] = "Fait"
fa["B1"] = "Vrai ?"
fa["C1"] = "Catégorie"
fa["D1"] = "Espèce établie"
for c in ("A1", "B1", "C1", "D1"):
    fa[c].fill = HDR_FILL
    fa[c].font = HDR_FONT
    fa[c].alignment = Alignment(horizontal="center")
FACT_END = len(facts) + 1   # 26 : dernière ligne de fait pré-rempli
BUF_END = 40                # lignes 27..40 : vides, prêtes pour de nouveaux faits
italic_grey = Font(name=FONT, size=10, italic=True, color="808080")
for i in range(2, BUF_END + 1):
    if i <= FACT_END:
        name, cat = facts[i - 2]
        fa.cell(i, 1, name).font = BASE_FONT
        fa.cell(i, 3, cat).font = italic_grey
        if cat.startswith("catégorie"):
            fa.cell(i, 1).fill = DERIV_FILL
        elif cat.startswith("espèce"):
            fa.cell(i, 1).fill = ESP_FILL
    fa.cell(i, 2).value = False
    fa.cell(i, 2).font = BASE_FONT
    fa.cell(i, 2).alignment = Alignment(horizontal="center")
    fa.cell(i, 4).value = f'=IF(AND($B{i}=TRUE,ISNUMBER(SEARCH("espèce",$C{i}))),$A{i},"")'
    fa.cell(i, 4).font = italic_grey
    for col in (1, 2, 3, 4):
        fa.cell(i, col).border = BORDER
fa.column_dimensions["A"].width = 24
fa.column_dimensions["B"].width = 9
fa.column_dimensions["C"].width = 22
fa.column_dimensions["D"].hidden = True
fa.freeze_panes = "A2"

# ---------- Base de règles ----------
br = wb.create_sheet("Base de règles")
rules = [
    ("R1", ["a du poil"], "mammifère"),
    ("R2", ["allaite ses petits"], "mammifère"),
    ("R3", ["a des plumes"], "oiseau"),
    ("R4", ["mammifère", "mange de la viande"], "carnivore"),
    ("R5", ["mammifère", "a des dents pointues", "a des griffes"], "carnivore"),
    ("R6", ["carnivore", "robe fauve", "taches sombres"], "guépard"),
    ("R7", ["carnivore", "robe fauve", "rayures noires"], "tigre"),
    ("R8", ["mammifère", "rumine", "long cou", "taches sombres"], "girafe"),
    ("R9", ["mammifère", "a des sabots", "rayures noires"], "zèbre"),
    ("R10", ["oiseau", "incapable de voler", "long cou"], "autruche"),
    ("R11", ["oiseau", "incapable de voler", "nage", "plumage noir et blanc"], "manchot"),
]
HDR_ROW = 2
FIRST = 3
MAXROW = 22  # règles en 3..13, lignes d'extension 14..22

# --- bandeau d'état (ligne 1) ---
br.merge_cells("A1:H1")
banner = br.cell(1, 1)
_esp = 'TEXTJOIN("",TRUE,Faits!$D$2:$D$40)'
banner.value = (
    '=IF(' + _esp + '<>"","🎯 Animal identifié : "&' + _esp + ','
    'IF(SUMPRODUCT(--(G3:G22=TRUE))>0,'
    '"▶ Continuez : déclenchez une règle surlignée en vert",'
    'IF(SUMPRODUCT(--(Faits!$B$2:$B$40=TRUE))>0,'
    '"⛔ Bloqué : aucune espèce trouvée — il faut ajouter une règle (étape Étendre)",'
    '"⏳ Cochez les caractéristiques du cas dans l’onglet Faits")))'
)
banner.font = Font(name=FONT, bold=True, size=12)
banner.alignment = Alignment(horizontal="center", vertical="center")
br.row_dimensions[1].height = 28
br.conditional_formatting.add(
    "A1:H1", FormulaRule(formula=['ISNUMBER(SEARCH("Animal",$A$1))'], fill=GREEN_OK))
br.conditional_formatting.add(
    "A1:H1", FormulaRule(formula=['ISNUMBER(SEARCH("Continuez",$A$1))'], fill=AMBER))
br.conditional_formatting.add(
    "A1:H1", FormulaRule(formula=['ISNUMBER(SEARCH("Bloqu",$A$1))'], fill=RED))

# --- en-têtes (ligne 2) ---
headers = ["Règle", "Condition 1", "Condition 2", "Condition 3", "Condition 4",
           "→ Conclusion", "Activable ? (calculé)", "Déjà conclue ? (calculé)"]
for j, h in enumerate(headers, start=1):
    c = br.cell(HDR_ROW, j, h)
    c.fill = HDR_FILL
    c.font = HDR_FONT
    c.alignment = Alignment(horizontal="center", wrap_text=True)

# --- règles (lignes 3..22) ---
for idx in range(FIRST, MAXROW + 1):
    r = idx - FIRST
    if r < len(rules):
        rid, conds, concl = rules[r]
        br.cell(idx, 1, rid)
        for k in range(4):
            br.cell(idx, 2 + k, conds[k] if k < len(conds) else None)
        br.cell(idx, 6, concl)
    g = (f'=IF($F{idx}="","",AND('
         f'IF($B{idx}="",TRUE,VLOOKUP($B{idx},Faits!$A:$B,2,FALSE)),'
         f'IF($C{idx}="",TRUE,VLOOKUP($C{idx},Faits!$A:$B,2,FALSE)),'
         f'IF($D{idx}="",TRUE,VLOOKUP($D{idx},Faits!$A:$B,2,FALSE)),'
         f'IF($E{idx}="",TRUE,VLOOKUP($E{idx},Faits!$A:$B,2,FALSE)),'
         f'NOT(VLOOKUP($F{idx},Faits!$A:$B,2,FALSE))))')
    br.cell(idx, 7, g)
    h = f'=IF($F{idx}="","",VLOOKUP($F{idx},Faits!$A:$B,2,FALSE))'
    br.cell(idx, 8, h)
    for j in range(1, 9):
        cell = br.cell(idx, j)
        cell.border = BORDER
        if j in (7, 8):
            cell.font = GREYFONT
            cell.fill = GREYFILL
            cell.alignment = Alignment(horizontal="center")
        else:
            cell.font = BASE_FONT

# menus déroulants pour conditions + conclusion
dv = DataValidation(type="list", formula1="Faits!$A$2:$A$40", allow_blank=True)
br.add_data_validation(dv)
dv.add(f"B{FIRST}:F{MAXROW}")

# surlignage vert des règles activables
br.conditional_formatting.add(
    f"A{FIRST}:H{MAXROW}", FormulaRule(formula=[f"$G{FIRST}=TRUE"], fill=GREEN))

widths = [7, 20, 20, 18, 22, 16, 13, 16]
for j, w in enumerate(widths, start=1):
    br.column_dimensions[chr(64 + j)].width = w
br.freeze_panes = "A3"

_out = os.path.join(_ROOT, "TN1-systeme-expert-animaux.xlsx")
wb.save(_out)
print("saved:", _out)
